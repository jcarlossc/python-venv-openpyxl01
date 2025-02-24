# Importa módulo Openpyxl
from openpyxl import load_workbook

# Carrega a planilha do diretório planilhas/
planilha_carregada = load_workbook("planilhas/planilha.xlsx")

# Selecionar a planilha ativa
planilha_ativa = planilha_carregada.active

# Percorre as células da planilha
for linha in planilha_ativa.iter_rows(values_only=True):
    print(linha)